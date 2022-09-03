# moduleImport
# autor: MolokovAlex
# lisence: GPL
# coding: utf-8

# модуль держатель элементов GUI для функций импорта-экспорта

# from sys import getsizeof
import openpyxl
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import END, N, NSEW, Menu
import tkinter.messagebox as mb
import sqlite3 as sqlite
import pandas as pd
import numpy as np


import modulAppGUI as mag
import skladConfig as scfg
import moduleDBClass as mdbc
import moduleSQLite as msql





class Window_Import_OR_Remove(tk.Toplevel):
    def __init__(self, parent, modeWindow, viewDB): 
           
        # визуальное отображение окна для БД компонентов или БД спецификаций :
        # viewWindow = DBC = DataBaseComponents
        # viewWindow = DBS = DataBaseSpecification
        # режим работы окна - или "перемещнние " или "импорт"
        # modeWindow = 'remove'
        # modeWindow == 'import
                                                                        
        super().__init__(parent)
        # self.geometry("1024x600")
        opts = { 'ipadx': 3, 'ipady': 3 , 'sticky': 'nswe' }

        if modeWindow == 'import':
            self.label_to = tk.Label(self, text="Импорт из XLSX файла ")
        elif modeWindow == 'remove':
            self.label_to = tk.Label(self, text="Куда переместить ? ")    
        self.label_to.grid(row=0, column=0, **opts)

        self.frame_tableTreeGroup = tk.LabelFrame(master=self, text="Группы", relief=tk.SUNKEN, borderwidth=3)   #, height=50)
        self.frame_tableTreeGroup.grid(row=1, column=0,columnspan=2, **opts)

        if modeWindow == 'import':
            self.btn_Remove = tk.Button(master=self, height=3, text="Импорт из XLSX", command=lambda vW = viewDB: self.press_Button_Import(viewWindow = vW))
        elif modeWindow == 'remove':
            self.btn_Remove = tk.Button(master=self, height=3, text="Сюда", command=self.fn_removComponent)
        self.btn_Remove.grid(row=2, column=0, **opts)
        self.btn_Cansel = tk.Button(master=self, height=3, text="Отмена", command=self.destroy)
        self.btn_Cansel.grid(row=2, column=1, **opts)

        self.treeGroup = ttk.Treeview(self.frame_tableTreeGroup, show="tree headings")#, columns=col) #self.columns)
        self.treeGroup.grid(row=0, column=1, rowspan=4, **opts)

        mag.Setting_TreeView(self.treeGroup, form = 'short')

        self.ysb = ttk.Scrollbar(self.frame_tableTreeGroup, orient=tk.VERTICAL, command=self.treeGroup.yview)
        self.treeGroup.configure(yscroll=self.ysb.set)
        self.ysb.grid(row=0, column=2, rowspan=4, **opts)

        self.treeGroup.bind('<<TreeviewSelect>>', self.on_select)

        self.new_parent = 0
        self.sel = 0
        
        # DBC = DataBaseComponents
        # DBS = DataBaseSpecification
        if viewDB == 'DBC':   dataFrame_in = scfg.df_DBC
        elif viewDB == 'DBS': dataFrame_in = scfg.df_DBS
        if not(dataFrame_in.empty):
            mag.viewTreeGroup(self.treeGroup)       # отобразим данные из памяти DataFrame в TreeGroup с родителями
        return None

    def press_Button_Import(self, viewWindow):
        if viewWindow == 'DBC':   dataFrame_in = scfg.df_DBC
        elif viewWindow == 'DBS': dataFrame_in = scfg.df_DBS
        Load_DBGroup_From_XLS_pandas(scfg.nameFile_importDBC_excel)
        mag.viewTreeGroup(self.treeGroup)       # отобразим данные из памяти DBGroup в Tree
        return None

    def open(self):
        self.grab_set()
        self.wait_window()
        usr = self.new_parent
        return usr
    
    def fn_removComponent(self):
        self.new_parent = self.sel[0]
        self.destroy()
        return None

    def on_select(self, event):
        self.sel = self.treeGroup.selection()
        return None















def import_Group_From_XLS_in_memory(nameFile_DBf, nameFile_DBgroup_import):
    """
    внесем новые данные из XLSX файла  - т.е. ИМПОРТ GROUP из XLSX файла 
    """
    DBGroupF = []
    DBGroup_import = []           # список БД по наименованию групп - временная, для загрузки из файла
    flag_copyTableInMemory = False
    sql_update_query_in_tableDBGroupComponent = 'UPDATE DBGroupComponent SET name_group = ? WHERE id_code_group = ?'
    sql_insert_data_in_tableDBGroupComponent = 'INSERT INTO DBGroupComponent (id_code_group, name_group) values(?, ?)' 

    
    # -------------------- считаем из XLS данные -------------------------------------------------------
    if foolproofDBGroupXLSX():  # пржд чем загружать данные - проанализирум файл на наличие ошибок в файле.
        DBGroup_import = Load_DBGroup_From_XLS_pandas(scfg.nameFile_importDBC_excel)     # создаем загрузку в базу DBGroup_import из файла xls
        # DBGroup_import = Load_DBGroup_From_XLS(nameFile_DBgroup_import)     # создаем загрузку в базу DBGroup_import из файла xls
        print ('БД в файле XLSX загружена в список DBGroup_import')

    else:
        print ('БД НЕ ЗАГРУЖЕНА!!!!')
    # print (f'в функции import_Group_From_XLS_in_ListDB размер DBGroup_import = {getsizeof( DBGroup_import)} бит')

    # ------------------- Считаем таблицы из файла БД_SQL и положим их в памятть --------------------------------------
    DBGroupF, flag_copyTableInMemory, flag_empty_Table = msql.copy_File_SQLDBGroupComponent_In_memory(scfg.nameFile_DB)
    if flag_empty_Table:
        message = "Файл БД SQLite не содержит записей в таблице DBGroupComponent\n Заполнить файл БД SQLite записями из XLS файла?" 
        if mb.askyesno(message=message):
            # заполняем файл БД SQLite записями из XLS файла
            connectionDBFile = sqlite.connect(nameFile_DBf)
            cursorDB = connectionDBFile.cursor()
            with connectionDBFile:
                for i in range (0, len(DBGroup_import), 1):
                    # data = []
                    # data.append(DBGroup_import[i].id_code_group)
                    # data.append(DBGroup_import[i].name_group)
                    cursorDB.execute(sql_insert_data_in_tableDBGroupComponent, (DBGroup_import[i].id_code_group_item, DBGroup_import[i].name_group))
                    connectionDBFile.commit()   
        # если нельзя заполнять файл БД SQLite записями из XLS файла - нчего не дделаем
        # else:
        #     ...
    # если файл БД SQLite содержит какие-то данные
    # ЕЩЕ раз Считаем таблицы из файла БД_SQL и положим их в памятть
    DBGroupF, flag_copyTableInMemory, flag_empty_Table = msql.copy_File_SQLDBGroupComponent_In_memory(scfg.nameFile_DB)
        #  теперь внесем изменения из DBGroup_import в DBGroup с одновременным внесением исправлений в файл БД через SQL
    connectionDBFile = sqlite.connect(nameFile_DBf)
    cursorDB = connectionDBFile.cursor()
    with connectionDBFile:
            for i in range (0, len(DBGroup_import), 1):
                flag_processing_item = True   # переменная факта обработки строки
                for j in range (0, len(DBGroupF), 1):
                    # code_group_db = str(DBGroupF[j].code_group[0]) + str(DBGroupF[j].code_group[1]) + str(DBGroupF[j].code_group[2]) + str(DBGroupF[j].code_group[3]) + str(DBGroupF[j].code_group[4]) + str(DBGroupF[j].code_group[5])
                    # code_group_import = str(DBGroup_import[i].code_group[0]) + str(DBGroup_import[i].code_group[1]) + str(DBGroup_import[i].code_group[2]) + str(DBGroup_import[i].code_group[3]) + str(DBGroup_import[i].code_group[4]) + str(DBGroup_import[i].code_group[5])
                    if DBGroup_import[i].id_code_group_item == DBGroupF[j].id_code_group_item:       #  если коды груп совпали
                        if DBGroupF[j].name_group != DBGroup_import[i].name_group:
                            # если наименования НЕсовпали, а коды совпали - вывести ошибку и вопрос
                            message = "Записи не равны по наименованию \n Code: " + DBGroup_import[i].id_code_group_item + "\n" + "Name in XLSX: " + str(DBGroup_import[i].name_group) + "\n" + "Name in DB: " + str(DBGroupF[j].name_group) + "\n\n" + "В XLSX правильно?"
                            if mb.askyesno(message=message):    #, parent=self):

                                DBGroupF[j].name_group = DBGroup_import[i].name_group      #  тогда запишем новое название из XLSX
                                p = cursorDB.execute("SELECT id_code_group FROM DBGroupComponent WHERE name_group = ?", DBGroupF[j].name_group)
                                a= p.id_code_group
                                column_values = (DBGroupF[j].name_group, p.id_code_group)
                                cursorDB.execute(sql_update_query_in_tableDBGroupComponent, column_values)
                                connectionDBFile.commit()
                                #app.text_box.insert(tk.END, "Строка успешно обновлена в файле БД"+"\n")
                                print("Строка успешно обновлена в файле БД")
                            flag_processing_item = False
                            break                               # мы поработали с близнецом - на новую иттерацию цикла DBGroup_temp
                        else:      # если наименования совпали, и коды совпали - ничего не делаем - мы нащли близнеца - на новую иттерацию цикла DBGroup_temp
                            flag_processing_item = False
                            break
                    else:       #  если коды груп НЕсовпали - ищем дальше совпадения - в  else ничего не делаем
                        flag_processing_item = True        #  подтверждаем что нет совпадения - нет обработки
                if flag_processing_item:        # если так и не нашлось ни одного совпадения кода группы - значит это новая группа (врученую добавленная в XLSX)
                    new_item_dbgoup = mdbc.Group(DBGroup_import[i].id_code_group_item, DBGroup_import[i].name_group)    # значит занесем ее в DBGroup
                    DBGroupF.append(new_item_dbgoup)
                    #print('Новая группа: ', DBGroup_import[i].code_group, DBGroup_import[i].name_group)
                    data = []
                    #gr = []
                    #data.append(getCode())                      # генерация уникального номера строки таблицы
                    #gr = DBGroup_import[i].code_group
                    # data.append(DBGroup_import[i].code_group[0])
                    # data.append(DBGroup_import[i].code_group[1])
                    # data.append(DBGroup_import[i].code_group[2])
                    # data.append(DBGroup_import[i].code_group[3])
                    # data.append(DBGroup_import[i].code_group[4])
                    # data.append(DBGroup_import[i].code_group[5])
                    # data.append(DBGroup_import[i].name_group)
                    cursorDB.execute(sql_insert_data_in_tableDBGroupComponent, (DBGroup_import[i].id_code_group_item, DBGroup_import[i].name_group))
                    connectionDBFile.commit() 
    # print (f'в функции import_Group_From_XLS_in_ListDB размер DBGroupF = {getsizeof( DBGroupF)} бит')
    return DBGroupF

def Load_DataFrameDBC_From_PickleFile():
    """
    функция считывания из файла Pickle базы DataFrame БД склада компонентов
    Вход:
    None
    Выход:
    None
    """
    scfg.df_DBC = pd.read_pickle(scfg.nameFile_DBC_pickle)
    return None

def Load_DataFrameDBS_From_PickleFile():
    """
    функция считывания из файла Pickle базы DataFrame БД спецификаций
    Вход:
    None
    Выход:
    None
    """
    scfg.df_DBS = pd.read_pickle(scfg.nameFile_DBS_pickle)
    return None

def Load_DataFrameDBI_From_PickleFile():
    """
    функция считывания из файла Pickle базы DataFrame БД спецификаций
    Вход:
    None
    Выход:
    None
    """
    scfg.df_DBI = pd.read_pickle(scfg.nameFile_DBI_pickle)
    return None

def Load_DataFrameDBE_From_PickleFile():
    """
    функция считывания из файла Pickle базы DataFrame БД спецификаций
    Вход:
    None
    Выход:
    None
    """
    scfg.df_DBE = pd.read_pickle(scfg.nameFile_DBE_pickle)
    return None

def Load_DataFrameDBCU_From_PickleFile():
    """
    функция считывания из файла Pickle базы DataFrame БД спецификаций
    Вход:
    None
    Выход:
    None
    """
    scfg.df_DBCU = pd.read_pickle(scfg.nameFile_DBCU_pickle)
    return None


def Load_DBGroup_From_XLS_pandas(nameFile: str): 
    """
    функция считывания из файла xls функциями Pandas
    Вход:
    nameFile -  имя основного файла, содержащего БД
    Выход:
    None
    """
    
    # with pd.ExcelFile(scfg.nameFile_importDB_excel) as xls:
    #     # scfg.df1 = pd.read_excel(xls, sheet_name=scfg.namesheet_DB_components)
    #     df1 = pd.read_excel(xls, sheet_name=scfg.namesheet_DB_components, usecols=scfg.import_columns_in_XLSX_file[1]) #, columns=scfg.columns_DB_components[1])
    #     df1.insert(0, "id_code_item", '')               # добавляем еще один столбец в позицию 0, т.е. в начало
    #     df2 = pd.read_excel(xls, sheet_name=scfg.namesheet_DB_components, usecols=scfg.import_columns_in_XLSX_file[2]) #, columns=scfg.columns_DB_components[2])
    #     df1 = pd.concat([df1, df2], ignore_index=True)
    # df1.to_excel("2_out.xlsx")

    # wb.save('temp02_out.xlsx')

    with pd.ExcelFile(scfg.nameFile_importDB_excel) as xlsFile:
        wb = openpyxl.load_workbook(xlsFile)
    ws = wb[scfg.namesheet_DB_components]
    # считае из свойств книги как визуально происходит групиировка
    # см. свойство "Данные - структура - Расположение итоговых данных - итоги в строках под данными"
    # "итоги в строках под данными" - галочка стоит - итоги раскрываются вверх (неудобно визуально) - wsprops.outlinePr.summaryBelow = True
    # "итоги в строках под данными" - галочка НЕстоит - итоги раскрываются вниз (удобно визуально) - wsprops.outlinePr.summaryBelow = False
    wsprops = ws.sheet_properties
    if not(wsprops.outlinePr.summaryBelow):
        prev_lvl = 0
        lvl = 0
        for itemRow in range(1,ws.max_row,1):
        # берем свойства строки
            row_dim = ws.row_dimensions[itemRow]
            # присвоим ячейки наименование
            lvl = row_dim.outline_level
            ws.cell(row=itemRow, column=4, value=lvl)
            if (prev_lvl+1) == lvl:
                cll_value = 'lvl0' + str(lvl)
                ws.cell(row=itemRow-1, column=5, value=cll_value)
                ws.cell(row=itemRow, column=5, value=cll_value)
            else: 
                cll_value = 'item_lvl' + str(lvl)
                ws.cell(row=itemRow, column=5, value=cll_value)
            prev_lvl = lvl
    else:
        prev_lvl = 0
        lvl = 0
        for itemRow in range(1,ws.max_row+2,1):
        # берем свойства строки
            row_dim = ws.row_dimensions[itemRow]
        # присвоим ячейки наименование
            lvl = row_dim.outline_level
            ws.cell(row=itemRow, column=4, value=lvl)
            if (prev_lvl-1) == lvl:
                cll_value = 'lvl0' + str(prev_lvl)
                # ws.cell(row=itemRow, column=5, value=cll_value)
                ws.cell(row=itemRow, column=2, value=cll_value)
            else:
                cll_value = 'item_lvl' + str(lvl)
                # ws.cell(row=itemRow, column=5, value=cll_value)
                ws.cell(row=itemRow, column=2, value=cll_value)

            prev_lvl = lvl


    wb.save('temp02_out.xlsx')

# создадим новый файл XLSX на базе импортируемого, но убрав все "украшательства" исходного файла и добавив новые колонки-поля
    with pd.ExcelFile(scfg.nameFile_importDB_excel) as xlsFile:
        # scfg.df1 = pd.read_excel(xls, sheet_name=scfg.namesheet_DB_components)
        # df1 = pd.DataFrame(data=None, columns=scfg.columns_DB_components)

        # считываем из Икселя столбец с названиями 'name'
        df1 = pd.read_excel(xlsFile, sheet_name=scfg.namesheet_DB_components, usecols=scfg.import_columns_in_XLSX_file[1], header=scfg.number_row_of_name_columns) 
        # счтитываем из Икселя столбец с количеством 'amount'
        df2 = pd.read_excel(xlsFile, sheet_name=scfg.namesheet_DB_components, usecols=scfg.import_columns_in_XLSX_file[2]) 
        df1 = pd.concat([df1, df2], axis=1)
        # считываем из Икселя столбец с служебное поле - буквенный код уровня вложенности родителя(группы) (поле только для группы) 'id_code_lvl'
        df2 = pd.read_excel(xlsFile, sheet_name=scfg.namesheet_DB_components, usecols=scfg.import_columns_in_XLSX_file[9])
        df1 = pd.concat([df1, df2], axis=1)

        # добавляем еще пустые столбцы в разные позиции
        df1.insert(0, "id_code_item", '')               
        df1.insert(3, 'code_units', '')
        df1.insert(4, 'min_rezerve', '')
        df1.insert(5, 'articul_1C', '')
        df1.insert(6, 'code_1C', '')
        df1.insert(7, 'name_1C', '')
        df1.insert(8, 'id_code_parent', '')
        df1.columns = scfg.columns_DB_components

        # df1.to_excel("2_out.xlsx")

    # columns = ['name', 'id_code_lvl', 'amount']
    # with pd.ExcelFile("2_out.xlsx") as xls:
    #         scfg.df1 = pd.read_excel(xls, sheet_name=scfg.namesheet_DB_components, names= columns)      #  скопируем данные из xlsx в DataFrame с заданием названия столбцов 
                             
    # приведем таблицу в порядок - заменим пустые строки в столбце "наимнование компонента" на NAN и удалим строки с NAN
    df1['name'].replace('', np.nan, inplace=True)       # 
    df1.dropna(subset=['name'], inplace=True)           # 
    print('1')

    df1.reset_index(drop=True, inplace=True)        # удалим пропущенные индексы
    for indx in df1.index:                              # присвоим каждому элементу - и ЭРЭ и группе ЭРЭ уникальные номера
        df1.loc[indx, 'id_code_item'] = mdbc.getCode()
        # print (df1.index.size)
        a = int(indx/df1.index.size*100)
        print (a," percent complete         \r",  end='')
    print ('\n')
    scfg.df_DBC = df1
    print('2')
    df1.to_excel("2_out.xlsx")    


# пройдем по всем индексам таблицы и напишем каждому кто является его родителем
#  такой странный порядок из-за того что Excel  делает группировку ВВЕРХ, т.е. название группы идет после всех элементов, входящих в группу
    
    listOfLevel = ['lvl01', 'lvl02', 'lvl03', 'lvl04', 'lvl05', 'lvl06']
    listOfIndexComponent = []
    listOfIndexParent = []

    for lvl in listOfLevel:
        listOfIndexComponent = []
        for indx in scfg.df_DBC.index:                              # пройдем по всем индексам таблицы
            res = False
            for k in listOfIndexParent:
                    if indx == k:
                        listOfIndexComponent = []
                        res = True
                        break
            if scfg.df_DBC.loc[indx, 'id_code_lvl'] == lvl:     
                id_group_parent =  scfg.df_DBC.loc[indx, 'id_code_item']       
                listOfIndexParent.append(indx)                           
                for i in listOfIndexComponent:                                 
                    scfg.df_DBC.loc[i, 'id_code_parent'] = id_group_parent      
                    listOfIndexComponent = [] 
            else:                                                       
                if not(res) :
                    listOfIndexComponent.append(indx)
    
    #дадим всем родителям уровня 'lvl01' родителя с кодом '1000000'
    for indx in scfg.df_DBC.index:                              # пройдем по всем индексам таблицы
        if scfg.df_DBC.loc[indx, 'id_code_lvl'] == listOfLevel[0]:
            scfg.df_DBC.loc[i, 'id_code_parent'] = '1000000'



    # print ('Save in file')
    # сохраняем изменненный DataFame в файл
    # Save_DataFrame_in_PickleFile()
    # Save_DataFrame_in_ExcelFile()
    # print ('ALL END')
    return None




def Load_DBGroup_From_XLS(nameFile: str):
    """
    функция считывания из файла xls функциями 
    Вход:
    nameFile -  имя основного файла, содержащего БД
    Выход:
    DBGroupFFX - список с объектми класса Group
    """
    DBGroupFFX = []
    codegroup = []
    namegroup = []
    codegroup_item = []
    namegroup_item = []
    
    # with pd.ExcelFile(nameFile) as xls:
    #     df1 = pd.read_excel(xls, "Sheet1")
        
    wb = openpyxl.load_workbook(nameFile)
    sheet = wb.active
    name_item = ''
    indexInFile = 1 
    s = sheet.cell(row=indexInFile, column=1).value         # если в первой колонке встретилось слово end - останавливаемся читать из файла
    flag_end_of_row, flag_pusto_in_str = Analiz_End_of_row(s)
    while not(flag_end_of_row) :                   # пока не встретилось слово end - считываем строкм
            if not(flag_pusto_in_str):
                codegroup_item = []
                namegroup_item = []
                codegroup = []
                namegroup = []
                codegroup_item.append(int(sheet.cell(row=indexInFile, column=1).value))     #[0] id_code_group_item
                codegroup_item.append(int(sheet.cell(row=indexInFile, column=2).value))           
                namegroup_item.append(str(sheet.cell(row=indexInFile, column=3).value))     #[0] это должен быть "Склад"  
                codegroup_item.append(int(sheet.cell(row=indexInFile, column=4).value))
                namegroup_item.append(str(sheet.cell(row=indexInFile, column=5).value))
                codegroup_item.append(int(sheet.cell(row=indexInFile, column=6).value))
                namegroup_item.append(str(sheet.cell(row=indexInFile, column=7).value))
                codegroup_item.append(int(sheet.cell(row=indexInFile, column=8).value))
                namegroup_item.append(str(sheet.cell(row=indexInFile, column=9).value))
                codegroup_item.append(int(sheet.cell(row=indexInFile, column=10).value))
                namegroup_item.append(str(sheet.cell(row=indexInFile, column=11).value))
                codegroup_item.append(int(sheet.cell(row=indexInFile, column=12).value))
                namegroup_item.append(str(sheet.cell(row=indexInFile, column=13).value))
        
                id_code_group_item = int(sheet.cell(row=indexInFile, column=1).value)
                code_group_item_YY, level_item_XX, code_group_parent_WW, level_parent_LL = mdbc.parsing_tree_parent(id_code_group_item)

                if level_item_XX == '1':
                    name_item = namegroup_item[0]
                elif level_item_XX == '2':
                    name_item = namegroup_item[1]
                elif level_item_XX == '3':
                    name_item = namegroup_item[2]
                elif level_item_XX == '4':
                    name_item = namegroup_item[3]
                elif level_item_XX == '5':
                    name_item = namegroup_item[4]
                elif level_item_XX == '6':
                    name_item = namegroup_item[5]
                item_dbgoup = mdbc.Group(id_code_group_item,  name_item)
                DBGroupFFX.append(item_dbgoup)   
            indexInFile = indexInFile + 1
            s = sheet.cell(row=indexInFile, column=1).value
            flag_end_of_row, flag_pusto_in_str = Analiz_End_of_row(s)
    #закрываем ее
    wb.close()
    #print ("OK file load XLS in DBComponent")
    # print (f'в функции LoadDBGroupFromFileXLS размер DBGroupF = {getsizeof( DBGroupFFX)} бит')
    return DBGroupFFX

# def Load_DBGroup_From_XLS(nameFile: str):
#     """
#     функция считывания из файла xls
#     Вход:
#     nameFile -  имя основного файла, содержащего БД
#     Выход:
#     DBGroupFFX - список с объектми класса Group
#     """
#     DBGroupFFX = []
#     codegroup = []
#     namegroup = []
#     codegroup_item = []
#     namegroup_item = []
    
#     wb = openpyxl.load_workbook(nameFile)
#     sheet = wb.active
#     name_item = ''
#     indexInFile = 1 
#     s = sheet.cell(row=indexInFile, column=1).value         # если в первой колонке встретилось слово end - останавливаемся читать из файла
#     flag_end_of_row, flag_pusto_in_str = Analiz_End_of_row(s)
#     while not(flag_end_of_row) :                   # пока не встретилось слово end - считываем строкм
#         if not(flag_pusto_in_str):
#             codegroup_item = []
#             namegroup_item = []
#             codegroup = []
#             namegroup = []
#             codegroup_item.append(int(sheet.cell(row=indexInFile, column=1).value))     #[0] id_code_group_item
#             codegroup_item.append(int(sheet.cell(row=indexInFile, column=2).value))           
#             namegroup_item.append(str(sheet.cell(row=indexInFile, column=3).value))     #[0] это должен быть "Склад"  
#             codegroup_item.append(int(sheet.cell(row=indexInFile, column=4).value))
#             namegroup_item.append(str(sheet.cell(row=indexInFile, column=5).value))
#             codegroup_item.append(int(sheet.cell(row=indexInFile, column=6).value))
#             namegroup_item.append(str(sheet.cell(row=indexInFile, column=7).value))
#             codegroup_item.append(int(sheet.cell(row=indexInFile, column=8).value))
#             namegroup_item.append(str(sheet.cell(row=indexInFile, column=9).value))
#             codegroup_item.append(int(sheet.cell(row=indexInFile, column=10).value))
#             namegroup_item.append(str(sheet.cell(row=indexInFile, column=11).value))
#             codegroup_item.append(int(sheet.cell(row=indexInFile, column=12).value))
#             namegroup_item.append(str(sheet.cell(row=indexInFile, column=13).value))
    
#             id_code_group_item = int(sheet.cell(row=indexInFile, column=1).value)
#             code_group_item_YY, level_item_XX, code_group_parent_WW, level_parent_LL = mdbc.parsing_tree_parent(id_code_group_item)

#             if level_item_XX == '1':
#                 name_item = namegroup_item[0]
#             elif level_item_XX == '2':
#                 name_item = namegroup_item[1]
#             elif level_item_XX == '3':
#                 name_item = namegroup_item[2]
#             elif level_item_XX == '4':
#                 name_item = namegroup_item[3]
#             elif level_item_XX == '5':
#                 name_item = namegroup_item[4]
#             elif level_item_XX == '6':
#                 name_item = namegroup_item[5]
#             item_dbgoup = mdbc.Group(id_code_group_item,  name_item)
#             DBGroupFFX.append(item_dbgoup)   
#         indexInFile = indexInFile + 1
#         s = sheet.cell(row=indexInFile, column=1).value
#         flag_end_of_row, flag_pusto_in_str = Analiz_End_of_row(s)
#     #закрываем ее
#     wb.close()
#     #print ("OK file load XLS in DBComponent")
#     print (f'в функции LoadDBGroupFromFileXLS размер DBGroupF = {getsizeof( DBGroupFFX)} бит')
#     return DBGroupFFX

    

def Analiz_Pusto_In_Str(instr: str) -> bool:
    """
    Анализирует строку на наличе пустой ячейки
    Вход:
    instring: str - входная строка
    Выход:
    flag_pustoF: bool -  флаг присуствия на наличе пустой ячейки
    """
    flag_pustoF = True
    
    if (instr == '' ) | (instr == '\n') | (instr == " ") | (instr == None):
        flag_pustoF = True
    else:
        flag_pustoF = False

    return flag_pustoF    

def Analiz_End_of_row(instring: int) -> bool:
    """
    Анализирует строку на наличе слова end в любом регистре
    Вход:
    instring: str - входная строка
    Выход:
    End_of_row: bool -  флаг присуствия слова end
    flag_pusto - флаг пустой ячейки (строки)
    """
    End_of_row = False
    flag_pusto = False
    instr = str(instring)
    flag_pusto = Analiz_Pusto_In_Str(instr)
    if not(flag_pusto):             # если первая ячейка НЕпустая - вероятно и вся строка пустая - 
        upper_instring = instr.upper()
        if upper_instring == 'END' :        # проверим строку на наличе слова end в любом регистре
            End_of_row = True
        else:
            End_of_row = False
    else:                           # если первая ячейка пустая - вероятно и вся строка пустая - не ищем слово end
        flag_pusto = True
    return End_of_row, flag_pusto

def foolproofDBGroupXLSX():
    return True