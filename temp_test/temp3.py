
# import openpyxl
# from ctypes import wstring_at
import pandas as pd
import numpy as np
import skladConfig as scfg
import moduleImport as mi
import moduleDBClass as mdbc


# https://russianblogs.com/article/2671820861/

import openpyxl #Подключаем библиотеку
# from openpyxl import Workbook
# from openpyxl.styles import PatternFill#Подключаем стили для ячеек
# from openpyxl.workbook import Workbook
# from openpyxl.styles import Font, Fill#Подключаем стили для текста
# from openpyxl.styles import colors#Подключаем цвета для текста и ячеек

# wb = openpyxl.load_workbook('temp02.xlsx')

# wb = openpyxl.load_workbook(scfg.nameFile_importDB_excel)
with pd.ExcelFile(scfg.nameFile_importDB_excel) as xlsFile:
    wb = openpyxl.load_workbook(xlsFile)
# wb = openpyxl.Workbook()
# ws = wb.active
ws = wb[scfg.namesheet_DB_components]
ws.title = 'RowColFold'
# for itemRow in range(1,20,1):
#     # row_dim = ws.row_dimensions[itemRow]
#     cll = ws.cell(row=itemRow, column=1, value='A3')

# считае из свойств книги как визуально происходит групиировка
# см. свойство "Данные - структура - Расположение итоговых данных - итоги в строках под данными"
# "итоги в строках под данными" - галочка стоит - итоги раскрываются вверх (неудобно визуально) - wsprops.outlinePr.summaryBelow = True
# "итоги в строках под данными" - галочка НЕстоит - итоги раскрываются вниз (удобно визуально) - wsprops.outlinePr.summaryBelow = False
wsprops = ws.sheet_properties
if not(wsprops.outlinePr.summaryBelow):
    prev_lvl = 0
    lvl = 0
    for itemRow in range(1,ws.max_row,1):
    # берем свойства строки
        row_dim = ws.row_dimensions[itemRow]
        # присвоим ячейки наименование
        lvl = row_dim.outline_level
        ws.cell(row=itemRow, column=4, value=lvl)
        if (prev_lvl+1) == lvl:
            cll_value = 'lvl0' + str(lvl)
            ws.cell(row=itemRow-1, column=5, value=cll_value)
            ws.cell(row=itemRow, column=5, value=cll_value)
        else: 
            cll_value = 'item_lvl' + str(lvl)
            ws.cell(row=itemRow, column=5, value=cll_value)
        prev_lvl = lvl
else:
    prev_lvl = 0
    lvl = 0
    for itemRow in range(1,ws.max_row+2,1):
    # берем свойства строки
        row_dim = ws.row_dimensions[itemRow]
    # присвоим ячейки наименование
        lvl = row_dim.outline_level
        ws.cell(row=itemRow, column=4, value=lvl)
        if (prev_lvl-1) == lvl:
            cll_value = 'lvl0' + str(prev_lvl)
            # ws.cell(row=itemRow, column=5, value=cll_value)
            ws.cell(row=itemRow, column=2, value=cll_value)
        else:
            cll_value = 'item_lvl' + str(lvl)
            # ws.cell(row=itemRow, column=5, value=cll_value)
            ws.cell(row=itemRow, column=2, value=cll_value)

        prev_lvl = lvl


wb.save('temp02_out.xlsx')
# свернем колонки с буквами с 'C' до 'F'
# ws.column_dimensions.group('C','F', hidden=True)
# свернем строки с номерами с 3 по 10
# ws.row_dimensions.group(3, 10, outline_level=1, hidden=True)
# теперь свернем строки с номерами с 7 по 10, с уровнем 2
# ws.row_dimensions.group(7,10, outline_level=2, hidden=True)
# cll_A3 = ws.cell(row=3, column=1, value='A3')
# cll_A7 = ws.cell(row=7, column=1, value='A7')
# for row in sorted(ws.row_dimensions):
    # col_dim = ws.column_dimensions[3]



        
    # ws.cell(row=itemRow, column=1, value=cll_value)
    # print(row_dim.outline_level, row_dim.hidden)
    


    # outline2=ws.dimensions[row].outline_level
    # print(row, ws.dimensions[row],  outline2 )

# wb.save('fold_row-col.xlsx')



# workbook = openpyxl.load_workbook(scfg.nameFile_importDB_excel)
# wb = openpyxl.Workbook() #Создали книгу
# work_sheet = wb.create_sheet(title='Test sheet') #Создали лист с названием и сделали его активным

# sheet = workbook[scfg.namesheet_DB_components]
 # Получить указанную ячейку
# cell = sheet.cell(row=1, column=1)

# пройдем по всему столбцу с названиями и там где название Bold жирное - это название группы - пометим 'LVL' в графе Количество
# maxRow = sheet.max_row()

# cells_range = sheet['A1':maxRow]
# for cells in cells_range:
#     for cell in cells:
#     	     # Получить значение атрибута
#         # print(cell.value)
#         # print(cell.font.bold)
#         if (cell.font.bold) and (cell.value != '')  :
#            r = cell.row
#            sheet.cell(row=r, column=2, value='LVL')

# workbook.save('03.xlsx')

# work_sheet['A1'] = 'Test text'
# work_sheet_a1 = work_sheet['A5']#Создали переменную, в которой содержится ячейка A1 с уже имеющимся текстом
# work_sheet_a1.font = Font(size=23, underline='single', color='FFBB00', bold=True, italic=True) #Применяем к тексту следующие параметры: размер — 23, подчеркивание, цвет = FFBB00 (цвет текста указывается в RGB), жирный, наклонный. Если нам не нужен жирный шрифт — используем конструкцию: bold=False. Аналогично действуем, если нам не нужен наклонный шрифт: italic=False.
# Важно: если есть необходимость — в стилях заложена возможность использования стандартных цветов, но код в этом случае будет выглядеть иначе:
# work_sheet_a1.font = Font(size=23, underline='single', color = colors.RED, bold=True, italic=True) #где color = colors.RED — прописанный в стилях цвет
# work_sheet_a1.fill = PatternFill(fill_type='solid', start_color='ff8327', end_color='ff8327')#Данный код позволяет делать оформление цветом ячейки



# создадим новый файл XLSX на базе импортируемого, но убрав все "украшательства" исходного файла и добавив новые колонки-поля
# with pd.ExcelFile(scfg.nameFile_importDB_excel) as xlsFile:
with pd.ExcelFile('temp02_out.xlsx') as xlsFile:
        # scfg.df1 = pd.read_excel(xls, sheet_name=scfg.namesheet_DB_components)
        # df1 = pd.DataFrame(data=None, columns=scfg.columns_DB_components)

        # считываем из Икселя столбец с названиями
        df1 = pd.read_excel(xlsFile, sheet_name='RowColFold', usecols=scfg.import_columns_in_XLSX_file[1], header=scfg.number_row_of_name_columns) 
        # счтитываем из Икселя столбец с количеством
        df2 = pd.read_excel(xlsFile, sheet_name='RowColFold', usecols=scfg.import_columns_in_XLSX_file[2]) 
        df1 = pd.concat([df1, df2], axis=1)
        # считываем из Икселя столбец с служебное поле - буквенный код уровня вложенности родителя(группы) (поле только для группы)
        df2 = pd.read_excel(xlsFile, sheet_name='RowColFold', usecols=scfg.import_columns_in_XLSX_file[9])
        df1 = pd.concat([df1, df2], axis=1)

        # добавляем еще пустые столбцы в разные позиции
        df1.insert(0, "id_code_item", '')               
        df1.insert(3, 'code_units', '')
        df1.insert(4, 'min_rezerve', '')
        df1.insert(5, 'articul_1C', '')
        df1.insert(6, 'code_1C', '')
        df1.insert(7, 'name_1C', '')
        df1.insert(8, 'id_code_parent', '')
        df1.columns = scfg.columns_DB_components

        # df1.to_excel("2_out.xlsx")

    # columns = ['name', 'id_code_lvl', 'amount']
    # with pd.ExcelFile("2_out.xlsx") as xls:
    #         scfg.df1 = pd.read_excel(xls, sheet_name=scfg.namesheet_DB_components, names= columns)      #  скопируем данные из xlsx в DataFrame с заданием названия столбцов 
                             
    # приведем таблицу в порядок - заменим пустые строки в столбце "наимнование компонента" на NAN и удалим строки с NAN
df1['name'].replace('', np.nan, inplace=True)       # 
df1.dropna(subset=['name'], inplace=True)           # 
print('1')

df1.reset_index(drop=True, inplace=True)        # удалим пропущенные индексы
for indx in df1.index:                              # присвоим каждому элементу - и ЭРЭ и группе ЭРЭ уникальные номера
        df1.loc[indx, 'id_code_item'] = mdbc.getCode()
scfg.df_DBC = df1
df1.to_excel("2_out.xlsx")
print('2')


# df1['amount'].replace('0', 'lvl', inplace=True)
# df1['amount'].replace(0, '', inplace=True)


# listOfIndexLVL = []

# пройдем по всем индексам таблицы и сделаем список индексов групп
# for indx in scfg.df1.index:
#     if (df1.loc[indx, 'id_code_lvl'] == 'LVL'):
#         listOfIndexLVL.append(indx)

# пройдем по всем индексам таблицы и напишем каждому кто является его родителем
#  такой странный порядок из-за того что Excel  делает группировку ВВЕРХ, т.е. название группы идет после всех элементов, входящих в группу
    
listOfLevel = ['lvl01', 'lvl02', 'lvl03', 'lvl04', 'lvl05', 'lvl06']
listOfIndexComponent = []
listOfIndexParent = []

for lvl in listOfLevel:
        listOfIndexComponent = []
        for indx in scfg.df_DBC.index:                              # пройдем по всем индексам таблицы
            res = False
            for k in listOfIndexParent:
                    if indx == k:
                        listOfIndexComponent = []
                        res = True
                        break
            if scfg.df_DBC.loc[indx, 'id_code_lvl'] == lvl:     
                id_group_parent =  scfg.df_DBC.loc[indx, 'id_code_item']       
                listOfIndexParent.append(indx)                           
                for i in listOfIndexComponent:                                 
                    scfg.df_DBC.loc[i, 'id_code_parent'] = id_group_parent      
                    listOfIndexComponent = [] 
            else:                                                       
                if not(res) :
                    listOfIndexComponent.append(indx)

df1.to_excel("04.xlsx")
print('3')